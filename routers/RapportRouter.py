import fastapi as _fast
import sqlalchemy.orm as _orm
import controllers.RapportController as _control
from config.connexion_db import get_db
from config.schemas import RapportSchema, RapportExcel
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import pandas as pd
from functions.exportExcelFunction import export_data_frame, get_data_pir_navire, get_data_pin_navire, get_pir_march
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy.sql import text
from uuid import uuid4
import shutil
from fastapi.responses import FileResponse
from config.models import Rapport


router = _fast.APIRouter(
    prefix="/api/rapport",
    tags=["API RAPPORT"]
)


# start row pour les données des PIR NAVIRE pour les touchers
start_row_pir_touch = 2
start_col_pir_touch = 7


@router.post("/download")
def download_rapport(req: RapportExcel, db: _orm.Session = _fast.Depends(get_db)):
    file_name = f"{req.date_debut.strftime('%d-%m-%Y')}-RAPPORT-{uuid4()}.xlsx"
    
    all_data = get_data_pir_navire(req.date_debut, req.date_milieu, req.date_fin, db)
    all_data_pin = get_data_pin_navire(req.date_debut, req.date_milieu, req.date_fin, db)
    march_data = get_pir_march(req.date_debut, req.date_milieu, req.date_fin, db)


    # pour les ports dans le sommaire
    data_s = {
        "Ports d'Intérêt Nationaux - PIN": [],
        "Ports d'Intérêt régionaux NAVIRES - PIR": [],
        "Ports d'Intérêt Régionaux BOTRY et autres Embarcations": []
    }
    request_ports = "SELECT ports.nom_port, ports.status FROM ports WHERE ports.apmf = true"

    request_ports_exec = db.execute(text(request_ports)).all()
    for result in request_ports_exec:
        if result[1] == "PIN":
            data_s["Ports d'Intérêt Nationaux - PIN"].append(result[0])
        elif result[1] == "PIR":
            data_s["Ports d'Intérêt régionaux NAVIRES - PIR"].append(result[0])
        else:
            data_s["Ports d'Intérêt Régionaux BOTRY et autres Embarcations"].append(result[0])

    max_length = max(
        len(data_s["Ports d'Intérêt Nationaux - PIN"]),
        len(data_s["Ports d'Intérêt régionaux NAVIRES - PIR"]),
        len(data_s["Ports d'Intérêt Régionaux BOTRY et autres Embarcations"])
    )    
    for key in data_s:
        while len(data_s[key]) < max_length:
            data_s[key].append('') 
    print(data_s)

    # Créez un DataFrame pour chaque table de données
    data = {
        'PORT': ["Toamasina", "Toamasina", "Toamasina",  "Toamasina", "Toamasina", "Toliara", "Toliara", "Toliara", "Toliara", "Toliara", "Mahajanga", "Mahajanga", "Mahajanga", "Mahajanga", "Mahajanga", "Antsiranana", "Antsiranana", "Antsiranana", "Antsiranana", "Antsiranana",],
        'TYPE DESSERTE': ["Cabotage Nationale", "Bornage", "CILR", "Bornage1", "Total", "Cabotage Nationale", "Bornage", "CILR", "Bornage1", "Total","Cabotage Nationale", "Bornage", "CILR", "Bornage1", "Total", "Cabotage Nationale", "Bornage", "CILR", "Bornage1", "Total"],
        '22 FEB 2023': ["1", "2", "3", "5", "6", "7", "5", "6", "7", "3", "3", "3", "3", "3", "3", "1","1", "2", "3", "5"],
        '22 MARS 2023': ["1", "2", "3", "5", "6", "7", "5", "6", "7", "3", "3", "3", "3", "3", "3", "1","1", "2", "3", "5"],
        '22 AVR 2023': ["1", "2", "3", "5", "6", "7", "5", "6", "7", "3", "3", "3", "3", "3", "3", "1","1", "2", "3", "5"],
    }
    
    data_pin = {
        'PORT': ["Toamasina", "Toamasina", "Toamasina",  "Toamasina", "Toamasina", "Toliara", "Toliara", "Toliara", "Toliara", "Toliara", "Mahajanga", "Mahajanga", "Mahajanga", "Mahajanga", "Mahajanga","Antsiranana", "Antsiranana", "Antsiranana",  "Antsiranana", "Antsiranana"],
        'TYPE DESSERTE': ["Cabotage Nationale", "Bornage", "Cabotage INTR", "CILR", "Total", "Cabotage Nationale", "Bornage", "Cabotage INTR", "CILR", "Total","Cabotage Nationale", "Bornage", "Cabotage INTR", "CILR", "Total", "Cabotage Nationale", "Bornage", "Cabotage INTR", "CILR", "Total"],
        '22 FEB 2023': ["1", "2", "3", "5", "6", "7", "5", "6", "7", "3", "3", "3", "3", "3", "3", "1", "2", "3", "5", "6"],
        '22 MARS 2023': ["1", "2", "3", "5", "6", "7", "5", "6", "7", "3", "3", "3", "3", "3", "3", "1", "2", "3", "5", "6"],
        '22 AVR 2023': ["1", "2", "3", "5", "6", "7", "5", "6", "7", "3", "3", "3", "3", "3", "3", "1", "2", "3", "5", "6"],
    }

    data_merch = {
        'PORT': ["Toamasina", "Toliara", "Antsiranana", "Mahajanga"],
        '22 FEB 2023': [1, 2, 2, 200],
        '22 MARS 2023': [1, 2, 2, 200],
        '22 AVR 2023': [1, 2, 2, 200],
    } 

    data_summary = data_s
    
    # le pas
    step_merge = len(list(set(data['TYPE DESSERTE'])))

    try:
        table_1 = pd.DataFrame(all_data)
        table_2 = pd.DataFrame(all_data_pin)
        table_merch = pd.DataFrame(data_merch)
        table_pin_2 = pd.DataFrame(data_pin)
        table_summary = pd.DataFrame(data_summary)
        
        with pd.ExcelWriter(f'{file_name}', engine='openpyxl') as writer:
            table_summary.to_excel(writer, sheet_name="SOMMAIRE", index=False, startcol=1, startrow=12, header=True)
            if all_data["PORT"]:
                table_1.to_excel(writer, sheet_name="PIR NAVIRE", index=False, startcol=0, startrow=2, header=True)
            if all_data_pin['PORT']:
                table_2.to_excel(writer, sheet_name="PIN NAVIRE", index=False, startcol=0, startrow=2, header=True)


    except Exception as e:
        print(e)
        raise _fast.HTTPException(_fast.status.HTTP_400_BAD_REQUEST, detail='SERVER_ERROR')
    
    try:
        wb = load_workbook(f'{file_name}')
        ws = wb["SOMMAIRE"]
        img = Image("assets/apmf.jpg")
        img.anchor = 'B2'
        img.width=110
        img.height=110
        ws.add_image(img)
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 24
        ws.column_dimensions['C'].width = 24
        ws.column_dimensions['D'].width = 24
        ws.column_dimensions['E'].width = 24
        ws.column_dimensions['F'].width = 16
        ws.column_dimensions['G'].width = 24
        ws.column_dimensions['H'].width = 24
        ws.column_dimensions['I'].width = 24
        ws.column_dimensions['J'].width = 24
        ws.column_dimensions['K'].width = 24
        ws.column_dimensions['L'].width = 24
        ws.sheet_view.showGridLines = False
        ws["D1"] = f"DEPUIS {req.date_debut.strftime('%d - %m - %Y')}"
        ws["D1"].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
        ws["D1"].font = Font(bold=True, size=18)
        ws["D1"].fill = PatternFill(start_color="DDDDEE", end_color="DDDDEE", fill_type="solid")

        ws["F2"] = "Année observé : "
        ws["F2"].font = Font(underline="singleAccounting", bold=True)
        ws["F4"].font = Font(underline="singleAccounting", bold=True)
        ws["F4"] = "Mois présentés : "
        ws["G2"] = "2023"
        ws["G4"] = f"{req.date_debut.strftime('%d - %m - %Y').capitalize()} à {req.date_fin.strftime('%d - %m - %Y').capitalize()}"

        thin_border = Side(border_style="thin", color="000000")
        border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

        ws.merge_cells('B10:D10')
        ws['B10'] = "RAPPORT MENSUEL DES STATISTIQUES PORTUAIRE"
        ws["B10"].font = Font(name='trebuchet ms', color='222222', size=24, bold=True)
        ws["B10"].alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        ws.row_dimensions[10].height = 60
        ws["B10"].fill = PatternFill(start_color='FFAA26', end_color='EE7806', fill_type='solid')

        for rows in ws.iter_rows(min_row=10, min_col=1):
            for cell in rows:
                cell.alignment = cell.alignment.copy(wrapText=True)

        for row in range(0, len(data_summary["Ports d'Intérêt Nationaux - PIN"])):
            ws[f"B{row+14}"].border = border
            ws[f"C{row+14}"].border = border
            ws[f"D{row+14}"].border = border

        wb.save(f'{file_name}')

        if all_data['PORT']:
            export_data_frame(
                ws_name='PIR NAVIRE',
                start_col=0,
                start_row=2,
                step=len(list(set(all_data['TYPE DESSERTE']))),
                data_length=len(all_data['PORT']),
                my_file=f'{file_name}',
                isMerchExist=True,
                data_length_march=len(data_merch["PORT"]),
                title="TOUCHER NAVIRE"
            )
        if all_data_pin["PORT"]:
            export_data_frame(
                ws_name='PIN NAVIRE',
                start_col=0,
                start_row=2,
                step=step_merge,
                data_length=len(all_data_pin['PORT']),
                my_file=f'{file_name}',
                title="TOUCHERS NAVIRES",
                is_first_data=False
            )
    except Exception as e:
        print(e)
        raise _fast.HTTPException(_fast.status.HTTP_400_BAD_REQUEST, detail='SERVER_ERROR')
    
    shutil.move(file_name, 'files/rapports/')
    try:
        new_rapport = Rapport(
            fichier=file_name,
            date_debut=req.date_debut,
            date_fin=req.date_fin
        )

        db.add(new_rapport)
        db.commit()
        db.refresh(new_rapport)
    except Exception as e:
        print(e)
        raise _fast.HTTPException(_fast.status.HTTP_400_BAD_REQUEST, detail='SAVE_FAIL')
    
    return "success"


@router.get('/api/download/{file}')
async def get_rapport(file: str, db: _orm.Session = _fast.Depends(get_db)):
    select_file = db.query(Rapport).filter(Rapport.fichier == file)
    if not select_file.first():
        raise _fast.HTTPException(_fast.status.HTTP_400_BAD_REQUEST, detail='ERROR')

    location = f'files/rapports/{file}'
    print(location)
    return FileResponse(location, filename=location)


@router.get('/all')
async def get_all_rapport(db: _orm.Session = _fast.Depends(get_db)):
    all_rapports = db.query(Rapport).all()

    return all_rapports


@router.delete('/delete/one/{id}')
async def delete_one(id: int, db: _orm.Session = _fast.Depends(get_db)):
    select_rapport = db.query(Rapport).filter(Rapport.id_rapport == id)
    if not select_rapport.first():
        raise _fast.HTTPException(_fast.status.HTTP_400_BAD_REQUEST, detail='ERROR')
    
    select_rapport.delete(synchronize_session=False)
    db.commit()

    
    return 'SUCCESS'
