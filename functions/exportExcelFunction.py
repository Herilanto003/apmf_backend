import pandas as pd
import fastapi as _fast
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from sqlalchemy.sql import text
import shutil

def export_data_frame(
    ws_name, start_col, start_row, step, data_length, my_file: str, bg_color: str = "ddeeff", data_length_march = 0 , isMerchExist = False, is_first_data=True, title = ""
):
    
    wb = load_workbook(my_file)
    ws = wb[ws_name]
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 28
    ws.column_dimensions['K'].width = 28
    ws.column_dimensions['L'].width = 28
    ws.column_dimensions['M'].width = 20

    ws.sheet_view.showGridLines = False

    thin_border = Side(border_style='thin', color='000000')
    border = Border(top=thin_border, left=thin_border, bottom=thin_border, right=thin_border)

    ws.merge_cells(start_column=start_col+1, start_row=start_row-1, end_column=start_col+4, end_row=start_row-1)
    ws.cell(row=start_row-1, column=start_col+1, value=title).font = Font(size=18, bold=True)
    ws.cell(row=start_row-1, column=start_col+1, value=title).fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
    
    # merging
    for row in range(start_row+2, data_length + 3, step):
        ws.merge_cells(start_column=start_col+1, end_column=start_col+1, start_row=row, end_row=row+(step-1))

    # bordure
    if is_first_data:
        for row in range(0, data_length):
            ws[f"A{row+4}"].alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
            ws[f"B{row+4}"].alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
            ws[f"C{row+4}"].alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
            ws[f"D{row+4}"].alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
            ws[f"E{row+4}"].alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
            if not isMerchExist:
                ws[f"A{row+4}"].border = border
                ws[f"B{row+4}"].border = border
                ws[f"C{row+4}"].border = border
                ws[f"D{row+4}"].border = border
                ws[f"E{row+4}"].border = border

            if ws[f"B{row+4}"].value == "Total":
                ws[f"B{row+4}"].fill = PatternFill(start_color='dddd99', end_color='dddd99', fill_type='solid')
                ws[f"C{row+4}"].fill = PatternFill(start_color='dddd99', end_color='dddd99', fill_type='solid')
                ws[f"D{row+4}"].fill = PatternFill(start_color='dddd99', end_color='dddd99', fill_type='solid')
                ws[f"E{row+4}"].fill = PatternFill(start_color='dddd99', end_color='dddd99', fill_type='solid')
    if not is_first_data: 
        for row in range(0, data_length):
            ws[f"A{row+4}"].border = border
            ws[f"B{row+4}"].border = border
            ws[f"C{row+4}"].border = border
            ws[f"D{row+4}"].border = border
            
            ws[f"A{row+4}"].alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
            ws[f"B{row+4}"].alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
            ws[f"C{row+4}"].alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
            ws[f"D{row+4}"].alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
            ws[f"E{row+4}"].alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)

            if ws[f"B{row+4}"].value == "Total":
                ws[f"A{row+4}"].fill = PatternFill(start_color='dddd99', end_color='dddd99', fill_type='solid')
                ws[f"B{row+4}"].fill = PatternFill(start_color='dddd99', end_color='dddd99', fill_type='solid')
                ws[f"C{row+4}"].fill = PatternFill(start_color='dddd99', end_color='dddd99', fill_type='solid')
                ws[f"D{row+4}"].fill = PatternFill(start_color='dddd99', end_color='dddd99', fill_type='solid')


    if isMerchExist:
        for row in range(0, data_length_march):
            ws[f"A{row+4}"].border = border
            ws[f"B{row+4}"].border = border
            ws[f"C{row+4}"].border = border
            ws[f"D{row+4}"].border = border
            ws[f"B{row+4}"].fill = PatternFill(start_color='FFFF77', end_color='FFFF77', fill_type='solid')
            ws[f"C{row+4}"].fill = PatternFill(start_color='FFFF77', end_color='FFFF77', fill_type='solid')
            ws[f"D{row+4}"].fill = PatternFill(start_color='FFFF77', end_color='FFFF77', fill_type='solid')


    # if isMerchExist:
    #     barchar = BarChart()
    #     data = Reference(ws, min_col=2, min_row=3, max_col=4)
    #     categorie = Reference(ws, min_col=1, max_col=1, min_row=4)
    #     print(data_length)

    #     barchar.add_data(data, titles_from_data=True)
    #     barchar.set_categories(categorie)
    #     ws.add_chart(barchar, f"A{data_length}")

    wb.save(my_file)


def get_data_pir_navire(date_d, date_m, date_f, db):
    request_pir_first = f"""
        SELECT p.nom_port, 
            SUM(CASE WHEN a.type_desserte = 'CN' THEN 1 ELSE 0 END) AS CN,
            SUM(CASE WHEN a.type_desserte = 'CILC' THEN 1 ELSE 0 END) AS CILC,
            SUM(CASE WHEN a.type_desserte = 'BO' THEN 1 ELSE 0 END) AS BO,
            SUM(CASE WHEN a.type_desserte = 'CIR' THEN 1 ELSE 0 END) AS CIR,
            SUM(CASE WHEN a.type_desserte IN ('CN', 'CILC', 'BO', 'CIR') THEN 1 ELSE 0 END) AS Total
        FROM ports p
        LEFT OUTER JOIN accostage a ON a.id_port_accoste = p.id_port
        WHERE p.apmf = true AND p.status = 'PIR'
        AND a.date_enreg BETWEEN '{date_d}' AND '{date_m}'
        GROUP BY p.nom_port
        UNION ALL
        SELECT p.nom_port, 
            0 AS CN,
            0 AS CILC,
            0 AS BO,
            0 AS CIR,
            0 AS Total
        FROM ports p
        WHERE p.apmf = true AND p.status = 'PIR'
        AND (p.id_port NOT IN (SELECT id_port_accoste FROM accostage WHERE date_enreg BETWEEN '{date_d}' AND '{date_m}'))
        ORDER BY nom_port;
    """
    request_pir_second = f"""
        SELECT p.nom_port, 
            SUM(CASE WHEN a.type_desserte = 'CN' THEN 1 ELSE 0 END) AS CN,
            SUM(CASE WHEN a.type_desserte = 'CILC' THEN 1 ELSE 0 END) AS CILC,
            SUM(CASE WHEN a.type_desserte = 'BO' THEN 1 ELSE 0 END) AS BO,
            SUM(CASE WHEN a.type_desserte = 'CIR' THEN 1 ELSE 0 END) AS CIR,
            SUM(CASE WHEN a.type_desserte IN ('CN', 'CILC', 'BO', 'CIR') THEN 1 ELSE 0 END) AS Total
        FROM ports p
        LEFT OUTER JOIN accostage a ON a.id_port_accoste = p.id_port
        WHERE p.apmf = true AND p.status = 'PIR'
        AND a.date_enreg BETWEEN '{date_m}' AND '{date_f}'
        GROUP BY p.nom_port
        UNION ALL
        SELECT p.nom_port, 
            0 AS CN,
            0 AS CILC,
            0 AS BO,
            0 AS CIR,
            0 AS Total
        FROM ports p
        WHERE p.apmf = true AND p.status = 'PIR'
        AND (p.id_port NOT IN (SELECT id_port_accoste FROM accostage WHERE date_enreg BETWEEN '{date_m}' AND '{date_f}'))
        ORDER BY nom_port;
    """
    ports = []
    ports1 = []
    type_desserte = []
    type_desserte1 = []
    all_type = []
    try:
        request_pir_first_exec = db.execute(text(request_pir_first)).all()
        for result in request_pir_first_exec:
            for type in range(0, len(result)):
                if (type+1) == len(result):
                    break
                type_desserte.append(result[type+1])
                ports.append(result[0])
            all_type.append("cabotage national")
            all_type.append("cilc")
            all_type.append("bo")
            all_type.append("cir")
            all_type.append("Total")
    except Exception as e:
        print(e)
        raise _fast.HTTPException(_fast.status.HTTP_400_BAD_REQUEST, detail='SERVER_ERROR')
    
    try:
        request_pir_second_exec = db.execute(text(request_pir_second)).all()
        for result in request_pir_second_exec:
            for type in range(0, len(result)):
                if (type+1) == len(result):
                    break
                type_desserte1.append(result[type+1])
                ports1.append(result[0])
    except Exception as e:
        print(e)
        raise _fast.HTTPException(_fast.status.HTTP_400_BAD_REQUEST, detail='SERVER_ERROR')

    return {
        'PORT': ports,
        'TYPE DESSERTE': all_type,
        f"{date_d.strftime('%d/%m/%Y')} - {date_m.strftime('%d/%m/%Y')}": type_desserte,
        f"{date_m.strftime('%d/%m/%Y')} - {date_f.strftime('%d/%m/%Y')}": type_desserte1,
    }

def get_data_pin_navire(date_d, date_m, date_f, db):
    request_pir_first = f"""
        SELECT p.nom_port, 
            SUM(CASE WHEN a.type_desserte = 'CN' THEN 1 ELSE 0 END) AS CN,
            SUM(CASE WHEN a.type_desserte = 'CILC' THEN 1 ELSE 0 END) AS CILC,
            SUM(CASE WHEN a.type_desserte = 'BO' THEN 1 ELSE 0 END) AS BO,
            SUM(CASE WHEN a.type_desserte = 'CIR' THEN 1 ELSE 0 END) AS CIR,
            SUM(CASE WHEN a.type_desserte IN ('CN', 'CILC', 'BO', 'CIR') THEN 1 ELSE 0 END) AS Total
        FROM ports p
        LEFT OUTER JOIN accostage a ON a.id_port_accoste = p.id_port
        WHERE p.apmf = true AND p.status = 'PIN'
        AND a.date_enreg BETWEEN '{date_d}' AND '{date_m}'
        GROUP BY p.nom_port
        UNION ALL
        SELECT p.nom_port, 
            0 AS CN,
            0 AS CILC,
            0 AS BO,
            0 AS CIR,
            0 AS Total
        FROM ports p
        WHERE p.apmf = true AND p.status = 'PIN'
        AND (p.id_port NOT IN (SELECT id_port_accoste FROM accostage WHERE date_enreg BETWEEN '{date_d}' AND '{date_m}'))
        ORDER BY nom_port;
    """
    request_pir_second = f"""
        SELECT p.nom_port, 
            SUM(CASE WHEN a.type_desserte = 'CN' THEN 1 ELSE 0 END) AS CN,
            SUM(CASE WHEN a.type_desserte = 'CILC' THEN 1 ELSE 0 END) AS CILC,
            SUM(CASE WHEN a.type_desserte = 'BO' THEN 1 ELSE 0 END) AS BO,
            SUM(CASE WHEN a.type_desserte = 'CIR' THEN 1 ELSE 0 END) AS CIR,
            SUM(CASE WHEN a.type_desserte IN ('CN', 'CILC', 'BO', 'CIR') THEN 1 ELSE 0 END) AS Total
        FROM ports p
        LEFT OUTER JOIN accostage a ON a.id_port_accoste = p.id_port
        WHERE p.apmf = true AND p.status = 'PIR'
        AND a.date_enreg BETWEEN '{date_m}' AND '{date_f}'
        GROUP BY p.nom_port
        UNION ALL
        SELECT p.nom_port, 
            0 AS CN,
            0 AS CILC,
            0 AS BO,
            0 AS CIR,
            0 AS Total
        FROM ports p
        WHERE p.apmf = true AND p.status = 'PIN'
        AND (p.id_port NOT IN (SELECT id_port_accoste FROM accostage WHERE date_enreg BETWEEN '{date_m}' AND '{date_f}'))
        ORDER BY nom_port;
    """
    ports = []
    ports1 = []
    type_desserte = []
    type_desserte1 = []
    all_type = []
    try:
        request_pir_first_exec = db.execute(text(request_pir_first)).all()
        for result in request_pir_first_exec:
            for type in range(0, len(result)):
                if (type+1) == len(result):
                    break
                type_desserte.append(result[type+1])
                ports.append(result[0])
            all_type.append("cabotage national")
            all_type.append("cilc")
            all_type.append("bo")
            all_type.append("cir")
            all_type.append("Total")
    except Exception as e:
        print(e)
        raise _fast.HTTPException(_fast.status.HTTP_400_BAD_REQUEST, detail='SERVER_ERROR')
    
    try:
        request_pir_second_exec = db.execute(text(request_pir_second)).all()
        for result in request_pir_second_exec:
            for type in range(0, len(result)):
                if (type+1) == len(result):
                    break
                type_desserte1.append(result[type+1])
                ports1.append(result[0])
    except Exception as e:
        print(e)
        raise _fast.HTTPException(_fast.status.HTTP_400_BAD_REQUEST, detail='SERVER_ERROR')
    return {
        'PORT': ports,
        'TYPE DESSERTE': all_type,
        f"{date_d.strftime('%d/%m/%Y')} - {date_m.strftime('%d/%m/%Y')}": type_desserte,
        f"{date_m.strftime('%d/%m/%Y')} - {date_f.strftime('%d/%m/%Y')}": type_desserte1,
    }


def get_pir_march(date_d, date_m, date_f, db):

    request_march = f"""
        SELECT sum(marchandises.tonnage) as somme, ports.nom_port
        FROM accostage, marchandises, ports
        WHERE accostage.id_accostage = marchandises.id_accostage_marchandise AND accostage.id_port_accoste = ports.id_port AND accostage.date_enreg BETWEEN '{date_d}' and '{date_m}'
        GROUP BY ports.nom_port
    """

    data_march = {
        'PORT': [],
        f"date1": [],
        f"date2": [],
    }
    request_march_exec = db.execute(text(request_march)).all()
    for result in request_march_exec:
        data_march["PORT"].append(result[1])
        data_march[f"date1"].append(result[0])

    return 'dd'